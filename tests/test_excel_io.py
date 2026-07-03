import openpyxl

from vmail.excel_io import read_emails, write_results


def _make_input(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Scraped_Email", "Brand_Name", "Website"])
    ws.append(["jane@example.com", "BrandA", "https://a.com"])
    ws.append([None, None, None])  # blank row must be skipped
    ws.append(["bob@test.com", "BrandB", None])
    path = tmp_path / "input.xlsx"
    wb.save(path)
    return str(path)


def test_read_detects_email_column_and_skips_blanks(tmp_path):
    rows, headers = read_emails(_make_input(tmp_path))
    assert headers == ["Scraped_Email", "Brand_Name", "Website"]
    assert len(rows) == 2
    assert rows[0]["_email"] == "jane@example.com"
    assert rows[1]["Brand_Name"] == "BrandB"


def test_write_results_two_sheets_and_colors(tmp_path):
    rows, headers = read_emails(_make_input(tmp_path))
    rows[0].update(Status="VALID", Safe_To_Send="YES", Confidence=95,
                   Reason="Mailbox exists", Provider="GOOGLE", Role_Account=False,
                   Disposable=False, Free_Provider=False, Catch_All="NOT_CATCH_ALL",
                   SMTP_Code=250, SMTP_Evidence="250 OK")
    rows[1].update(Status="INVALID", Safe_To_Send="NO", Confidence=99,
                   Reason="Dead domain", Provider="OTHER", Role_Account=False,
                   Disposable=False, Free_Provider=False, Catch_All=None,
                   SMTP_Code=None, SMTP_Evidence=None)
    out = str(tmp_path / "out.xlsx")
    write_results(out, rows, headers)

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Results", "Summary"]
    ws = wb["Results"]
    header_row = [c.value for c in ws[1]]
    assert header_row[:3] == ["Scraped_Email", "Brand_Name", "Website"]
    assert "Status" in header_row and "Safe_To_Send" in header_row
    status_col = header_row.index("Status") + 1
    assert ws.cell(row=2, column=status_col).value == "VALID"
    assert ws.cell(row=2, column=1).fill.start_color.rgb.endswith("C6EFCE")
    assert ws.cell(row=3, column=1).fill.start_color.rgb.endswith("FFC7CE")

    summary = wb["Summary"]
    values = [tuple(r) for r in summary.iter_rows(values_only=True)]
    assert ("VALID", 1) in values
    assert ("INVALID", 1) in values


def test_write_results_risky_and_temporary_fill_colors(tmp_path):
    rows, headers = read_emails(_make_input(tmp_path))
    rows[0].update(Status="UNVERIFIABLE", Safe_To_Send="UNKNOWN", Confidence=0,
                   Reason="Could not verify", Provider="OTHER", Role_Account=False,
                   Disposable=False, Free_Provider=False, Catch_All=None,
                   SMTP_Code=None, SMTP_Evidence=None)
    rows[1].update(Status="TEMPORARY", Safe_To_Send="NO", Confidence=0,
                   Reason="Temporary error", Provider="OTHER", Role_Account=False,
                   Disposable=False, Free_Provider=False, Catch_All=None,
                   SMTP_Code=None, SMTP_Evidence=None)
    out = str(tmp_path / "out.xlsx")
    write_results(out, rows, headers)

    wb = openpyxl.load_workbook(out)
    ws = wb["Results"]
    assert ws.cell(row=2, column=1).fill.start_color.rgb.endswith("FFEB9C")
    assert ws.cell(row=3, column=1).fill.start_color.rgb.endswith("D9D9D9")

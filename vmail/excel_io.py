"""Excel input/output. Never writes over the input file."""
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill

RESULT_COLUMNS = [
    "Status", "Safe_To_Send", "Confidence", "Reason", "Provider",
    "Role_Account", "Disposable", "Free_Provider", "Catch_All",
    "SMTP_Code", "SMTP_Evidence",
]

_FILLS = {
    "VALID": PatternFill("solid", start_color="C6EFCE"),
    "INVALID": PatternFill("solid", start_color="FFC7CE"),
    "VALID_RISKY": PatternFill("solid", start_color="FFEB9C"),
    "UNVERIFIABLE": PatternFill("solid", start_color="FFEB9C"),
    "BLOCKED_BY_SERVER": PatternFill("solid", start_color="FFEB9C"),
    "TEMPORARY": PatternFill("solid", start_color="D9D9D9"),
    "ERROR": PatternFill("solid", start_color="D9D9D9"),
}


def read_emails(path: str) -> tuple[list[dict], list[str]]:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(rows_iter)]
    email_idx = next(
        (i for i, h in enumerate(headers) if "email" in h.lower() or "mail" in h.lower()),
        0,
    )
    rows = []
    for raw in rows_iter:
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        row = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}
        email = raw[email_idx] if email_idx < len(raw) else None
        if email is None or str(email).strip() == "":
            continue
        row["_email"] = str(email).strip()
        rows.append(row)
    wb.close()
    return rows, headers


def write_results(path: str, rows: list[dict], headers: list[str]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    all_headers = headers + RESULT_COLUMNS
    ws.append(all_headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(h) for h in all_headers])
        fill = _FILLS.get(row.get("Status", ""))
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    summary = wb.create_sheet("Summary")
    summary.append(["Status", "Count"])
    for status, count in Counter(r.get("Status") for r in rows).most_common():
        summary.append([status, count])
    summary.append([None, None])
    summary.append(["Safe_To_Send", "Count"])
    for bucket, count in Counter(r.get("Safe_To_Send") for r in rows).most_common():
        summary.append([bucket, count])
    summary.append([None, None])
    summary.append(["Top Domains", "Count"])
    domains = Counter(
        str(r.get("_email", "")).rsplit("@", 1)[-1].lower()
        for r in rows if "@" in str(r.get("_email", ""))
    )
    for domain, count in domains.most_common(15):
        summary.append([domain, count])

    wb.save(path)
